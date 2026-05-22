"""
excel_updater.py
----------------
Writes matched payment amounts into the master ledger (גביה 2026 sheet).

Ledger layout (confirmed from file inspection):
  - Sheet:      "גביה 2026"
  - Row 1:      Headers — col B = דירה, cols E–P = ינואר…דצמבר
  - Rows 2–61:  Apartment data (apts 1–60, one per row)
  - Row 62:     סה"כ  — SUM formulas        ← NEVER TOUCH
  - Row 63:     יעד   — target formulas      ← NEVER TOUCH
  - Row 64:     פער   — gap % formulas       ← NEVER TOUCH

Safety rules (hard-coded, not configurable):
  1. Never write to rows >= 62.
  2. Never overwrite a formula cell (starts with "=").
  3. Always create a timestamped backup before first write.
"""

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Hebrew month → MM mapping ─────────────────────────────────────────────────
_HEB_TO_MM = {
    "ינואר": "01", "פברואר": "02", "מרץ": "03",   "אפריל": "04",
    "מאי":   "05", "יוני":   "06", "יולי": "07",   "אוגוסט": "08",
    "ספטמבר":"09", "אוקטובר":"10", "נובמבר":"11",  "דצמבר":  "12",
}
_MM_TO_HEB = {v: k for k, v in _HEB_TO_MM.items()}

SHEET_NAME  = "גביה 2026"
HEADER_ROW  = 1
APT_COL     = 2   # column B
FIRST_DATA_ROW  = 2
LAST_DATA_ROW   = 61   # row 62+ are formulas — hard stop


def _month_to_mm(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.month:02d}"
    s = str(value).strip()
    import re as _re
    s_clean = _re.sub(r"[\s\d]+$", "", s).strip()
    if s_clean in _HEB_TO_MM:
        return _HEB_TO_MM[s_clean]
    if s in _HEB_TO_MM:
        return _HEB_TO_MM[s]
    import re
    m = re.fullmatch(r"(\d{1,2})[/\-\.](\d{4})", s)
    if m:
        return f"{int(m.group(1)):02d}"
    m = re.fullmatch(r"(\d{4})[/\-\.](\d{1,2})", s)
    if m:
        return f"{int(m.group(2)):02d}"
    try:
        n = int(s)
        if 1 <= n <= 12:
            return f"{n:02d}"
    except ValueError:
        pass
    return None


def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


# ── Petty cash (קופה קטנה) ────────────────────────────────────────────────────
PETTY_CASH_COL     = 17   # column Q
PETTY_CASH_DEFAULT = 500


class LedgerReader:
    """
    Lightweight, read-only snapshot of the גביה 2026 sheet.
    """
    def __init__(self, ledger_path: str):
        wb = openpyxl.load_workbook(str(ledger_path), data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{SHEET_NAME}' not found")
        ws = wb[SHEET_NAME]

        self._month_col: dict[str, int] = {}
        for cell in ws[HEADER_ROW]:
            mm = _month_to_mm(cell.value)
            if mm:
                self._month_col[mm] = cell.column

        self._apt_row: dict[int, int] = {}
        self._vals:    dict[tuple, object] = {}
        for row_idx in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            apt_val = ws.cell(row=row_idx, column=APT_COL).value
            try:
                self._apt_row[int(apt_val)] = row_idx
            except (TypeError, ValueError):
                pass
            for col_idx in list(self._month_col.values()) + [PETTY_CASH_COL]:
                self._vals[(row_idx, col_idx)] = ws.cell(row=row_idx, column=col_idx).value
        wb.close()

    def read_payment(self, apartment: int, month: str):
        mm  = _month_to_mm(month)
        row = self._apt_row.get(int(apartment))
        col = self._month_col.get(mm) if mm else None
        if row is None or col is None:
            return None
        return self._vals.get((row, col))

    def read_petty(self, apartment: int):
        row = self._apt_row.get(int(apartment))
        if row is None:
            return None
        return self._vals.get((row, PETTY_CASH_COL))


class LedgerUpdater:
    def __init__(self, ledger_path: str):
        self.path = Path(ledger_path)

        # ════════════ SMART BACKUP SYSTEM ════════════
        backup_dir = Path(r"C:\Users\danie\OneDrive\backup")
        backup_dir.mkdir(parents=True, exist_ok=True)

        today_str = datetime.now().strftime("%Y_%m_%d")
        base_name = self.path.stem
        backup_filename = f"{base_name}_{today_str}{self.path.suffix}"
        backup_path = backup_dir / backup_filename

        if not backup_path.exists():
            shutil.copy2(self.path, backup_path)
            print(f"✓ Backup saved: {backup_filename}")
        else:
            print("✓ Backup for today already exists. Skipped to preserve first state.")

        all_backups = sorted(backup_dir.glob(f"{base_name}_*{self.path.suffix}"))
        if len(all_backups) > 5:
            for old_backup in all_backups[:-5]:
                old_backup.unlink()
                print(f"✓ Deleted old backup: {old_backup.name}")
        # ═════════════════════════════════════════════

        self.wb = openpyxl.load_workbook(str(self.path))
        if SHEET_NAME not in self.wb.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. Available: {self.wb.sheetnames}"
            )
        self.ws = self.wb[SHEET_NAME]

        self._month_col: dict[str, int] = {}
        for cell in self.ws[HEADER_ROW]:
            mm = _month_to_mm(cell.value)
            if mm:
                self._month_col[mm] = cell.column

        self._apt_row: dict[int, int] = {}
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            val = self.ws.cell(row=row, column=APT_COL).value
            try:
                apt = int(val)
                self._apt_row[apt] = row
            except (TypeError, ValueError):
                pass

        print(
            f"✓ Ledger ready: {len(self._month_col)} month cols, "
            f"{len(self._apt_row)} apartment rows"
        )

    def write_payment(
        self, apartment: int, month: str, amount: float,
        overwrite: bool = False, mode: str = "write"
    ) -> tuple[bool, str]:
        mm  = _month_to_mm(month)
        row = self._apt_row.get(int(apartment))
        col = self._month_col.get(mm) if mm else None

        if row is None:
            return False, f"Apt {apartment} not found in ledger"
        if col is None:
            return False, f"Month '{month}' not found in ledger"
        if row > LAST_DATA_ROW:
            return False, f"Row {row} is in the formula section — refused"

        cell = self.ws.cell(row=row, column=col)
        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — skipped"

        col_letter = get_column_letter(col)

        if mode == "add":
            try:
                existing = float(cell.value) if cell.value not in (None, "", 0) else 0.0
            except (TypeError, ValueError):
                existing = 0.0
            new_total = existing + amount
            cell.value = new_total
            return True, (
                f"✓ {col_letter}{row}  דירה {apartment}  "
                f"₪{existing:.2f} + ₪{amount:.2f} = ₪{new_total:.2f}"
            )

        existing = cell.value
        if existing not in (None, "", 0) and not overwrite:
            return False, (
                f"{cell.coordinate}: already has {existing} — "
                f"skipped (pass overwrite=True to replace)"
            )
        cell.value = amount
        return True, f"✓ {col_letter}{row}  דירה {apartment}  ₪{amount:.2f}"

    def write_petty_cash(
        self, apartment: int, amount: float, mode: str = "write",
    ) -> tuple[bool, str]:
        row = self._apt_row.get(int(apartment))
        if row is None:
            return False, f"Apt {apartment} not found in ledger"
        if row > LAST_DATA_ROW:
            return False, f"Row {row} is in the formula section — refused"

        cell = self.ws.cell(row=row, column=PETTY_CASH_COL)
        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — skipped"

        col_letter = get_column_letter(PETTY_CASH_COL)

        if mode == "add":
            try:
                existing = float(cell.value) if cell.value not in (None, "", 0) else 0.0
            except (TypeError, ValueError):
                existing = 0.0
            new_total = existing + amount
            cell.value = new_total
            return True, (
                f"✓ {col_letter}{row}  דירה {apartment}  קופה קטנה  "
                f"₪{existing:.2f} + ₪{amount:.2f} = ₪{new_total:.2f}"
            )

        existing = cell.value
        if existing not in (None, "", 0):
            return False, f"{cell.coordinate}: already has {existing} — skipped"
        cell.value = amount
        return True, f"✓ {col_letter}{row}  דירה {apartment}  קופה קטנה  ₪{amount:.2f}"

    def save(self, output_path: str | None = None) -> str:
        out = Path(output_path) if output_path else self.path
        self.wb.save(str(out))
        print(f"✓ Saved → {out.name}")
        return str(out)

    @property
    def available_months(self) -> list[str]:
        return sorted(self._month_col.keys())

    @property
    def available_apts(self) -> list[int]:
        return sorted(self._apt_row.keys())


def apply_matches(
    matched: list[dict],
    ledger_path: str,
    payment_month: str,
    overwrite: bool = False,
) -> tuple[list[dict], list[dict]]:
    updater = LedgerUpdater(ledger_path)
    written, skipped = [], []

    for result in matched:
        tenant = result.get("tenant")
        if not tenant:
            skipped.append({**result, "write_message": "no tenant resolved"})
            continue

        if result.get("needs_manual"):
            skipped.append({**result, "write_message": "סכום שונה מהצפוי – הועבר לבדיקה ידנית"})
            continue

        amount   = result["transaction"]["amount"]
        tx_month = result["transaction"].get("payment_month") or payment_month
        success, msg = updater.write_payment(
            tenant["apartment"], tx_month, amount, overwrite=overwrite
        )
        entry = {**result, "write_message": msg}
        (written if success else skipped).append(entry)

    updater.save()
    print(f"Written: {len(written)}  |  Skipped: {len(skipped)}")
    return written, skipped


# ══════════════════════════════════════════════════════════════════════════════
# Raw transactions tab
# ══════════════════════════════════════════════════════════════════════════════

# Tab name prefix, e.g. "תנועות 05/2026"
_RAW_TAB_COLS = [
    "תאריך", "כניסה/יציאה", "תיאור", "אסמכתא", "סכום (₪)",
    "שם שולח / ישות", "דירה", "דייר", "קטגוריה",
    "חודש תשלום", "שיטת התאמה", "הערות", "תאור מורחב",
]

_FILL_CREDIT  = PatternFill("solid", fgColor="D6F0D6")   # light green  – income
_FILL_DEBIT   = PatternFill("solid", fgColor="FCE4D6")   # light salmon – expense
_FILL_HEADER  = PatternFill("solid", fgColor="2E4057")   # dark blue    – header
_FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
_FONT_NORMAL  = Font(size=10)
_ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=False)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center")


def write_raw_transactions_tab(
    all_rows: list[dict],
    all_match_results: list[dict],
    ledger_path: str,
    payment_month: str,
) -> tuple[int, str]:
    """
    Create (or replace) a worksheet named "תנועות MM/YYYY" containing every
    row from the bank statement — credits and debits — enriched with tenant
    and category info where available.

    Parameters
    ----------
    all_rows          : output of parse_all_rows()
    all_match_results : matched + unmatched lists from match_transactions(),
                        concatenated — used to enrich credit rows with tenant info
    ledger_path       : path to the master ledger workbook
    payment_month     : "MM/YYYY" string used for the sheet name

    Returns
    -------
    (row_count, sheet_name)
    """
    # ── Build enrichment lookup: raw_detail → match result ───────────────────
    # raw_detail is unique per bank row even when אסמכתא repeats (99012 etc.)
    match_by_detail: dict[str, dict] = {}
    for result in all_match_results:
        tx = result.get("transaction") or {}
        key = str(tx.get("raw_detail", "")).strip()
        if key:
            match_by_detail[key] = result

    # ── Resolve sheet name ────────────────────────────────────────────────────
    mm   = _month_to_mm(payment_month) or "??"
    year = str(payment_month).split("/")[-1] if "/" in str(payment_month) else "2026"
    heb_name = _MM_TO_HEB.get(mm, mm)
    sheet_name = f"תנועות {heb_name} {year}"

    # ── Open workbook and (re)create sheet ───────────────────────────────────
    wb = openpyxl.load_workbook(str(ledger_path))
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.rightToLeft = True   # RTL

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(_RAW_TAB_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    # Column widths
    col_widths = [12, 10, 18, 12, 10, 22, 7, 20, 18, 12, 14, 20, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_row = 2
    for r in all_rows:
        is_credit = r["direction"] == "credit"
        fill      = _FILL_CREDIT if is_credit else _FILL_DEBIT
        dir_label = "הכנסה ✅" if is_credit else "הוצאה 💸"

        # Enrich credit rows with tenant match info
        apt_num      = ""
        tenant_name  = ""
        match_method = ""
        note         = ""

        if is_credit:
            key    = str(r.get("raw_detail", "")).strip()
            result = match_by_detail.get(key)
            if result:
                tenant = result.get("tenant")
                if tenant:
                    apt_num     = str(tenant.get("apartment", ""))
                    tenant_name = tenant.get("tenant_name", "")
                match_method = result.get("match_method", "")
                if result.get("amount_mismatch"):
                    note = f"⚠️ צפוי ₪{(result.get('tenant') or {}).get('monthly_fee','?')}"
                if result.get("needs_manual"):
                    note = (note + " | בדיקה ידנית").lstrip(" | ")
            elif r.get("apt_hint"):
                apt_num = str(r["apt_hint"])

        # Sender name for credits, entity name for debits
        display_name = r.get("sender_name") or r.get("entity_name") or ""

        values = [
            r.get("date", ""),
            dir_label,
            r.get("description", ""),
            r.get("ref", ""),
            r.get("amount", 0),
            display_name,
            apt_num,
            tenant_name,
            r.get("category") or "",
            r.get("payment_month") or "",
            match_method,
            note,
            r.get("raw_detail") or "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=data_row, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = _FONT_NORMAL
            cell.alignment = _ALIGN_RIGHT

        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(str(ledger_path))
    rows_written = data_row - 2
    print(f"✓ Raw tab '{sheet_name}': {rows_written} rows written")
    return rows_written, sheet_name


# ══════════════════════════════════════════════════════════════════════════════
# Expense sheet updater  (הוצאות 2026)
# ══════════════════════════════════════════════════════════════════════════════

EXPENSE_SHEET_NAME  = "הוצאות 2026"
EXPENSE_HEADER_ROW  = 1
EXPENSE_CAT_COL     = 1
EXPENSE_FIRST_DATA  = 2
EXPENSE_LAST_DATA   = 14
EXPENSE_PROTECTED_COLS = {14, 15}
EXPENSE_COMMENT_ROW = 23


class ExpenseUpdater:
    def __init__(self, ledger_path: str):
        self.path = Path(ledger_path)

        self.wb = openpyxl.load_workbook(str(self.path))
        if EXPENSE_SHEET_NAME not in self.wb.sheetnames:
            raise ValueError(
                f"Sheet '{EXPENSE_SHEET_NAME}' not found. "
                f"Available: {self.wb.sheetnames}"
            )
        self.ws = self.wb[EXPENSE_SHEET_NAME]

        self._month_col: dict[str, int] = {}
        for cell in self.ws[EXPENSE_HEADER_ROW]:
            if cell.column in EXPENSE_PROTECTED_COLS:
                continue
            if cell.column <= 1:
                continue
            mm = _month_to_mm(cell.value)
            if mm:
                self._month_col[mm] = cell.column

        self._cat_row: dict[str, int] = {}
        for row in range(EXPENSE_FIRST_DATA, EXPENSE_LAST_DATA + 1):
            val = self.ws.cell(row=row, column=EXPENSE_CAT_COL).value
            if val:
                self._cat_row[str(val).strip()] = row

        print(
            f"✓ Expense sheet ready: {len(self._month_col)} month cols, "
            f"{len(self._cat_row)} category rows"
        )

    @property
    def available_categories(self) -> list[str]:
        return sorted(self._cat_row.keys())

    def write_expense(
        self, category: str, month: str, amount: float, overwrite: bool = False,
    ) -> tuple[bool, str]:
        mm  = _month_to_mm(month)
        row = self._cat_row.get(str(category).strip())
        col = self._month_col.get(mm) if mm else None

        if row is None:
            return False, f"Category '{category}' not found in expense sheet"
        if col is None:
            return False, f"Month '{month}' not found in expense sheet"
        if row > EXPENSE_LAST_DATA:
            return False, f"Row {row} is outside the expense data range — refused"
        if col in EXPENSE_PROTECTED_COLS:
            return False, f"Column {col} is a totals/average column — refused"

        cell = self.ws.cell(row=row, column=col)
        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — skipped"

        existing = cell.value
        if existing not in (None, "", 0) and not overwrite:
            return False, (
                f"{cell.coordinate}: already has {existing} — "
                f"skipped (enable 'דרוס ערכים קיימים' to replace)"
            )

        cell.value = amount
        col_letter = get_column_letter(col)
        return True, f"✓ {col_letter}{row}  {category}  ₪{amount:.2f}"

    def add_to_expense(self, category: str, month: str, amount: float) -> tuple[bool, str]:
        mm  = _month_to_mm(month)
        row = self._cat_row.get(str(category).strip())
        col = self._month_col.get(mm) if mm else None

        if row is None:
            return False, f"Category '{category}' not found in expense sheet"
        if col is None:
            return False, f"Month '{month}' not found in expense sheet"
        if row > EXPENSE_LAST_DATA:
            return False, f"Row {row} is outside the expense data range — refused"
        if col in EXPENSE_PROTECTED_COLS:
            return False, f"Column {col} is a totals/average column — refused"

        cell = self.ws.cell(row=row, column=col)
        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — skipped"

        try:
            existing = float(cell.value) if cell.value not in (None, "") else 0.0
        except (TypeError, ValueError):
            existing = 0.0

        new_total = existing + amount
        cell.value = new_total
        col_letter = get_column_letter(col)
        return True, (
            f"✓ {col_letter}{row}  {category}  "
            f"₪{existing:.2f} + ₪{amount:.2f} = ₪{new_total:.2f}"
        )

    def write_comment(self, month: str, text: str) -> tuple[bool, str]:
        mm  = _month_to_mm(month)
        col = self._month_col.get(mm) if mm else None
        if col is None:
            return False, f"Month '{month}' not found — comment not written"
        if col in EXPENSE_PROTECTED_COLS:
            return False, "Protected column — refused"

        cell = self.ws.cell(row=EXPENSE_COMMENT_ROW, column=col)
        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — comment skipped"

        existing = str(cell.value).strip() if cell.value not in (None, "") else ""
        cell.value = f"{existing} | {text}" if existing else text
        col_letter = get_column_letter(col)
        return True, f"✓ comment → {col_letter}{EXPENSE_COMMENT_ROW}: {text}"

    def save(self, output_path: str | None = None) -> str:
        out = Path(output_path) if output_path else self.path
        self.wb.save(str(out))
        print(f"✓ Expense sheet saved → {out.name}")
        return str(out)


def apply_expense_matches(
    matched_debits: list[dict],
    ledger_path: str,
    payment_month: str,
    overwrite: bool = False,
) -> tuple[list[dict], list[dict]]:
    from collections import defaultdict

    grouped: dict[str, float] = defaultdict(float)
    group_entries: dict[str, list[dict]] = defaultdict(list)
    for d in matched_debits:
        cat = d.get("category") or ""
        if cat:
            grouped[cat] += d["amount"]
            group_entries[cat].append(d)

    updater = ExpenseUpdater(ledger_path)
    written, skipped = [], []

    for cat, total_amount in grouped.items():
        success, msg = updater.write_expense(cat, payment_month, total_amount, overwrite=overwrite)
        entry = {
            "category":      cat,
            "amount":        total_amount,
            "count":         len(group_entries[cat]),
            "transactions":  group_entries[cat],
            "write_message": msg,
        }
        (written if success else skipped).append(entry)

    updater.save()
    print(f"Expense written: {len(written)}  |  Skipped: {len(skipped)}")
    return written, skipped


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "האגמית7_כספים_2026.xlsx"
    u = LedgerUpdater(path)
    print("Months:", u.available_months)
    print("Apts  :", u.available_apts[:10], "…")
# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 – Financial Dashboard additions
# Append everything below to the bottom of excel_updater.py
# ══════════════════════════════════════════════════════════════════════════════

# ── HOA constants ─────────────────────────────────────────────────────────────
PAYING_APTS  = 56    # apartments that pay monthly fee (4 committee apts exempt)
MONTHLY_FEE  = 350   # ₪ per month per paying apartment
TOTAL_APTS   = 60    # all apartments including committee
PETTY_AMOUNT = 500   # ₪ per apartment per year for capital fund

# ── Wishlist tab constants ─────────────────────────────────────────────────────
WISHLIST_SHEET      = "רשימת פרויקטים"
WISHLIST_COLS       = ["שם פרויקט", "עלות משוערת (₪)", "סטטוס", "הערות"]
WISHLIST_STATUS_OPT = ["מתוכנן", "בביצוע", "הושלם"]

# ── Expense sheet row/column constants used by dashboard ─────────────────────
EXPENSE_TOTAL_COL           = 15   # Column O  – סה"כ שנתי (formula, read-only)
EXPENSE_AVG_COL             = 14   # Column N  – live monthly average (formula, read-only)
EXPENSE_FIXED_ROWS          = list(range(2, 13))   # rows 2–12  Bucket 1 fixed costs
EXPENSE_KELALI_ROW          = 13   # row 13  כללי – variable maintenance
EXPENSE_TOTAL_OPERATING_ROW = 16   # row 16  הוצאות שוטף (formula, read-only)
EXPENSE_PETTY_TOTAL_ROW     = 17   # row 17  הוצאות לא שוטף (formula, read-only)
EXPENSE_COMMENT_ROW_DASH    = 23   # row 23  free-text comment per month

# ── קופה קטנה tab constants ───────────────────────────────────────────────────
PETTY_CASH_SHEET  = "קופה קטנה"
PETTY_SPENT_ROWS  = [3, 4, 5]    # rows 3–5: projects spent this year


class LedgerDashboardReader:
    """
    Read-only snapshot of all financial data needed for Tab 6.
    Every public attribute documents its ledger source via a companion _src dict.

    Sources:
      fixed_rows          → הוצאות 2026, rows 2–12, cols B–L (monthly), M (total), N (avg)
      kelali              → הוצאות 2026, row 13
      petty_collected     → גביה 2026, column Q (PETTY_CASH_COL = 17), rows 2–61
      payment_grid        → גביה 2026, month columns E–P, rows 2–61
      petty_spent_projects→ קופה קטנה, rows 3–5
    """

    @staticmethod
    def _safe_float(val) -> float:
        if val is None or val == "" or val == 0:
            return 0.0
        try:
            return float(str(val).replace(",", ""))
        except (TypeError, ValueError):
            return 0.0

    def _read_expense_row(self, ws, row_idx: int) -> dict:
        category = ws.cell(row=row_idx, column=EXPENSE_CAT_COL).value
        monthly: dict[str, float | None] = {}
        for mm, col in self._exp_month_cols.items():
            raw = ws.cell(row=row_idx, column=col).value
            monthly[mm] = self._safe_float(raw) if raw not in (None, "", 0) else None

        total = self._safe_float(ws.cell(row=row_idx, column=EXPENSE_TOTAL_COL).value)
        avg   = self._safe_float(ws.cell(row=row_idx, column=EXPENSE_AVG_COL).value)
        actuals = [v for v in monthly.values() if v is not None and v > 0]
        return {
            "category":     str(category).strip() if category else "",
            "monthly":      monthly,
            "total":        total,          # col O – הוצאות 2026 · עמודה O (סה"כ שנתי)
            "avg":          avg,            # col N – הוצאות 2026 · עמודה N
            "min":          min(actuals) if actuals else 0.0,
            "max":          max(actuals) if actuals else 0.0,
            "months_count": len(actuals),
            "row":          row_idx,
        }

    def __init__(self, ledger_path: str):
        self.path = Path(ledger_path)
        wb = openpyxl.load_workbook(str(ledger_path), data_only=True)

        # ── Expense sheet ────────────────────────────────────────────────────
        if EXPENSE_SHEET_NAME not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{EXPENSE_SHEET_NAME}' not found")
        ws_exp = wb[EXPENSE_SHEET_NAME]

        # Month-column mapping (skip avg col N and annual-total col O)
        self._exp_month_cols: dict[str, int] = {}
        for cell in ws_exp[EXPENSE_HEADER_ROW]:
            if cell.column in (EXPENSE_TOTAL_COL, EXPENSE_AVG_COL, EXPENSE_CAT_COL):
                continue
            mm = _month_to_mm(cell.value)
            if mm:
                self._exp_month_cols[mm] = cell.column

        # Fixed cost rows 2–12
        self.fixed_rows: list[dict] = []
        for row_idx in EXPENSE_FIXED_ROWS:
            rd = self._read_expense_row(ws_exp, row_idx)
            if rd["category"]:
                self.fixed_rows.append(rd)

        # כללי row 13
        self.kelali = self._read_expense_row(ws_exp, EXPENSE_KELALI_ROW)

        # Read-only summary rows (formulas, data_only=True gives computed values)
        self.row_total_operating = self._safe_float(   # row 16 col O
            ws_exp.cell(row=EXPENSE_TOTAL_OPERATING_ROW, column=EXPENSE_TOTAL_COL).value
        )
        self.row_petty_total = self._safe_float(       # row 17 col O
            ws_exp.cell(row=EXPENSE_PETTY_TOTAL_ROW, column=EXPENSE_TOTAL_COL).value
        )

        # Comment row 23 per month
        self.kelali_comments: dict[str, str] = {}
        for mm, col in self._exp_month_cols.items():
            v = ws_exp.cell(row=EXPENSE_COMMENT_ROW_DASH, column=col).value
            if v:
                self.kelali_comments[mm] = str(v).strip()

        # ── גביה 2026 – monthly payments + column Q petty cash ───────────────
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{SHEET_NAME}' not found")
        ws_ledger = wb[SHEET_NAME]

        ledger_month_cols: dict[str, int] = {}
        for cell in ws_ledger[HEADER_ROW]:
            mm = _month_to_mm(cell.value)
            if mm:
                ledger_month_cols[mm] = cell.column

        # apt → ₪ collected in column Q (PETTY_CASH_COL = 17)
        self.petty_collected: dict[int, float] = {}
        # apt → mm → ₪ paid
        self.payment_grid: dict[int, dict[str, float]] = {}

        for row_idx in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            apt_val = ws_ledger.cell(row=row_idx, column=APT_COL).value
            try:
                apt = int(apt_val)
            except (TypeError, ValueError):
                continue

            petty_val = ws_ledger.cell(row=row_idx, column=PETTY_CASH_COL).value
            self.petty_collected[apt] = self._safe_float(petty_val)

            self.payment_grid[apt] = {}
            for mm, col in ledger_month_cols.items():
                v = ws_ledger.cell(row=row_idx, column=col).value
                self.payment_grid[apt][mm] = self._safe_float(v)

        # ── קופה קטנה tab – projects spent this year ─────────────────────────
        self.petty_spent_projects: list[dict] = []
        if PETTY_CASH_SHEET in wb.sheetnames:
            ws_petty = wb[PETTY_CASH_SHEET]
            for row_idx in PETTY_SPENT_ROWS:
                name = ws_petty.cell(row=row_idx, column=1).value
                if not name:
                    continue
                total_spent = 0.0
                for col_idx in range(2, 14):  # columns B–M
                    total_spent += self._safe_float(
                        ws_petty.cell(row=row_idx, column=col_idx).value
                    )
                self.petty_spent_projects.append({
                    "name":   str(name).strip(),
                    "total":  total_spent,
                    "source": f"קופה קטנה · שורה {row_idx}",
                })

        wb.close()
        self._compute()

    def _compute(self):
        """Pre-compute all derived metrics once."""

        # ── Bucket 1: Fixed costs ─────────────────────────────────────────
        self.fixed_ytd         = sum(r["total"] for r in self.fixed_rows)
        self.fixed_avg_monthly = sum(r["avg"]   for r in self.fixed_rows)
        self.fixed_min_monthly = sum(r["min"]   for r in self.fixed_rows)
        self.fixed_max_monthly = sum(r["max"]   for r in self.fixed_rows)
        self.fixed_proj_annual = self.fixed_avg_monthly * 12
        self.fixed_proj_min    = self.fixed_min_monthly * 12
        self.fixed_proj_max    = self.fixed_max_monthly * 12

        # Months that have any expense entry (to gauge average confidence)
        months_seen: set[str] = set()
        for row in self.fixed_rows:
            for mm, v in row["monthly"].items():
                if v is not None and v > 0:
                    months_seen.add(mm)
        self.months_with_data = len(months_seen)

        # ── Income potential & YTD collected ─────────────────────────────
        # Source: tenants.csv constant + גביה 2026 all month columns
        self.annual_income_potential = PAYING_APTS * MONTHLY_FEE * 12  # ₪235,200
        self.ytd_collected = sum(
            sum(months.values()) for months in self.payment_grid.values()
        )

        # ── Coverage ratio: avg monthly income ÷ avg monthly fixed costs ──
        months_elapsed = max(self.months_with_data, 1)
        self.avg_monthly_collected = self.ytd_collected / months_elapsed
        self.coverage_ratio = (
            self.avg_monthly_collected / self.fixed_avg_monthly
            if self.fixed_avg_monthly > 0 else 0.0
        )

        # ── Bucket 2: Variable maintenance pool ───────────────────────────
        # Source: annual_income_potential – projected fixed annual – כללי YTD
        self.kelali_ytd  = self.kelali["total"]
        self.kelali_avg  = self.kelali["avg"]
        self.bucket2_remaining = (
            self.annual_income_potential
            - self.fixed_proj_annual
            - self.kelali_ytd
        )
        self.bucket2_monthly_rate = self.kelali_ytd / months_elapsed if months_elapsed > 0 else 0.0
        self.bucket2_months_runway = (
            self.bucket2_remaining / self.bucket2_monthly_rate
            if self.bucket2_monthly_rate > 0 else 99.0
        )

        # ── Bucket 3: Capital fund ────────────────────────────────────────
        # Collected: גביה 2026, col Q, rows 2–61 (all 60 apartments)
        # Spent: קופה קטנה tab, rows 3–5
        self.petty_total_collected = sum(self.petty_collected.values())
        self.petty_total_spent     = sum(p["total"] for p in self.petty_spent_projects)
        self.petty_available       = self.petty_total_collected - self.petty_total_spent
        self.petty_potential       = TOTAL_APTS * PETTY_AMOUNT   # ₪30,000
        self.petty_collection_pct  = (
            self.petty_total_collected / self.petty_potential * 100
            if self.petty_potential > 0 else 0.0
        )
        self.petty_apts_paid = sum(
            1 for v in self.petty_collected.values() if v >= PETTY_AMOUNT
        )


# ══════════════════════════════════════════════════════════════════════════════
# Wishlist manager
# ══════════════════════════════════════════════════════════════════════════════

_WISH_PLAN_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow  – planned
_WISH_WIP_FILL  = PatternFill("solid", fgColor="D6F0D6")   # green   – in progress
_WISH_DONE_FILL = PatternFill("solid", fgColor="E8E8E8")   # grey    – completed


class WishlistManager:
    """
    Read and write the capital projects wishlist tab (רשימת פרויקטים).

    Tab is created automatically on first use if missing.
    Completed items are stored and displayed but NOT deducted from the
    available balance (they are already captured as actual spend in קופה קטנה).
    """

    def __init__(self, ledger_path: str):
        self.path = Path(ledger_path)
        self._ensure_sheet()

    def _ensure_sheet(self):
        wb = openpyxl.load_workbook(str(self.path))
        if WISHLIST_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(WISHLIST_SHEET)
            ws.sheet_view.rightToLeft = True
            for col_idx, header in enumerate(WISHLIST_COLS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font      = _FONT_HEADER
                cell.fill      = _FILL_HEADER
                cell.alignment = _ALIGN_HEADER
            for i, w in enumerate([32, 16, 14, 42], 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"
            wb.save(str(self.path))
            print(f"✓ Created wishlist sheet '{WISHLIST_SHEET}'")
        wb.close()

    def read(self) -> list[dict]:
        wb = openpyxl.load_workbook(str(self.path), data_only=True)
        if WISHLIST_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[WISHLIST_SHEET]
        items = []
        for row_idx in range(2, (ws.max_row or 1) + 2):
            name = ws.cell(row=row_idx, column=1).value
            if not name:
                continue
            try:
                cost = float(ws.cell(row=row_idx, column=2).value or 0)
            except (TypeError, ValueError):
                cost = 0.0
            status_raw = str(ws.cell(row=row_idx, column=3).value or "מתוכנן").strip()
            status = status_raw if status_raw in WISHLIST_STATUS_OPT else "מתוכנן"
            notes  = str(ws.cell(row=row_idx, column=4).value or "").strip()
            items.append({
                "name":   str(name).strip(),
                "cost":   cost,
                "status": status,
                "notes":  notes,
                "row":    row_idx,
            })
        wb.close()
        return items

    def save(self, items: list[dict]) -> None:
        wb = openpyxl.load_workbook(str(self.path))
        if WISHLIST_SHEET not in wb.sheetnames:
            wb.close()
            self._ensure_sheet()
            wb = openpyxl.load_workbook(str(self.path))
        ws = wb[WISHLIST_SHEET]

        fill_map = {
            "מתוכנן": _WISH_PLAN_FILL,
            "בביצוע": _WISH_WIP_FILL,
            "הושלם":  _WISH_DONE_FILL,
        }

        # Clear all old data rows first
        max_r = max(ws.max_row or 1, len(items) + 1)
        for row_idx in range(2, max_r + 2):
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell.fill  = PatternFill()

        # Write items with colour coding
        for i, item in enumerate(items, 2):
            fill = fill_map.get(item.get("status", "מתוכנן"), _WISH_PLAN_FILL)
            vals = [
                item.get("name", ""),
                item.get("cost", 0.0),
                item.get("status", "מתוכנן"),
                item.get("notes", ""),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell           = ws.cell(row=i, column=col_idx, value=val)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font      = Font(size=10)

        wb.save(str(self.path))
        print(f"✓ Wishlist saved: {len(items)} projects → {WISHLIST_SHEET}")
