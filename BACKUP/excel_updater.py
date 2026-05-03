"""
excel_updater.py
----------------
Writes matched payment amounts into the master ledger (גביה 2026 sheet).

Ledger layout (confirmed from file inspection):
  - Sheet:      "גביה 2026"
  - Row 1:      Headers — col B = דירה, cols E–P = ינואר…דצמבר
  - Rows 2–61:  Apartment data (apts 1–60, one per row)
  - Row 62:     סה"כ  — SUM formulas        ← NEVER TOUCH
  - Row 63:     יעד   — target formulas      ← NEVER TOUCH
  - Row 64:     פער   — gap % formulas       ← NEVER TOUCH

Safety rules (hard-coded, not configurable):
  1. Never write to rows >= 62.
  2. Never overwrite a formula cell (starts with "=").
  3. Always create a timestamped backup before first write.
"""

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ── Hebrew month → MM mapping ─────────────────────────────────────────────────
_HEB_TO_MM = {
    "ינואר": "01", "פברואר": "02", "מרץ": "03",   "אפריל": "04",
    "מאי":   "05", "יוני":   "06", "יולי": "07",   "אוגוסט": "08",
    "ספטמבר":"09", "אוקטובר":"10", "נובמבר":"11",  "דצמבר":  "12",
}

SHEET_NAME  = "גביה 2026"
HEADER_ROW  = 1
APT_COL     = 2   # column B
FIRST_DATA_ROW  = 2
LAST_DATA_ROW   = 61   # row 62+ are formulas — hard stop


def _month_to_mm(value) -> str | None:
    """
    Normalise any month representation to a 2-digit month string "MM".
    Accepts: Hebrew names, "04/2026", "2026-04", datetime objects, int 1-12.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.month:02d}"
    s = str(value).strip()
    # Hebrew name (exact key lookup)
    if s in _HEB_TO_MM:
        return _HEB_TO_MM[s]
    # MM/YYYY or M/YYYY
    import re
    m = re.fullmatch(r"(\d{1,2})[/\-\.](\d{4})", s)
    if m:
        return f"{int(m.group(1)):02d}"
    m = re.fullmatch(r"(\d{4})[/\-\.](\d{1,2})", s)
    if m:
        return f"{int(m.group(2)):02d}"
    # Plain integer 1-12
    try:
        n = int(s)
        if 1 <= n <= 12:
            return f"{n:02d}"
    except ValueError:
        pass
    return None


def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


class LedgerUpdater:
    def __init__(self, ledger_path: str):
        self.path = Path(ledger_path)

        # ════════════ SMART BACKUP SYSTEM ════════════
        backup_dir = Path(r"C:\Users\danie\OneDrive\backup")
        backup_dir.mkdir(parents=True, exist_ok=True)

        # 1. Format today's date (e.g., 2026_05_02)
        today_str = datetime.now().strftime("%Y_%m_%d")
        base_name = self.path.stem
        backup_filename = f"{base_name}_{today_str}{self.path.suffix}"
        backup_path = backup_dir / backup_filename

        # 2. Save UNTOUCHED file ONLY if today's backup doesn't exist
        if not backup_path.exists():
            shutil.copy2(self.path, backup_path)
            print(f"✓ Backup saved: {backup_filename}")
        else:
            print("✓ Backup for today already exists. Skipped to preserve first state.")

        # 3. Enforce 5-day retention policy
        # glob finds all backups for this file, sorted() orders them chronologically by YYYY_MM_DD
        all_backups = sorted(backup_dir.glob(f"{base_name}_*{self.path.suffix}"))
        
        if len(all_backups) > 5:
            # Delete the oldest files until only 5 remain
            for old_backup in all_backups[:-5]:
                old_backup.unlink()
                print(f"✓ Deleted old backup: {old_backup.name}")
        # ═════════════════════════════════════════════

        self.wb = openpyxl.load_workbook(str(self.path))
        if SHEET_NAME not in self.wb.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. Available: {self.wb.sheetnames}"
            )
        self.ws = self.wb[SHEET_NAME]

        # Build month column map: "MM" → col index
        self._month_col: dict[str, int] = {}
        for cell in self.ws[HEADER_ROW]:
            mm = _month_to_mm(cell.value)
            if mm:
                self._month_col[mm] = cell.column

        # Build apartment row map: apt_int → row index
        self._apt_row: dict[int, int] = {}
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            val = self.ws.cell(row=row, column=APT_COL).value
            try:
                apt = int(val)
                self._apt_row[apt] = row
            except (TypeError, ValueError):
                pass

        print(
            f"✓ Ledger ready: {len(self._month_col)} month cols, "
            f"{len(self._apt_row)} apartment rows"
        )
        self.path = Path(ledger_path)

        # Backup on first open
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = self.path.with_suffix(f".backup_{ts}.xlsx")
        shutil.copy2(self.path, backup)
        print(f"✓ Backup: {backup.name}")

        self.wb = openpyxl.load_workbook(str(self.path))
        if SHEET_NAME not in self.wb.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. Available: {self.wb.sheetnames}"
            )
        self.ws = self.wb[SHEET_NAME]

        # Build month column map: "MM" → col index
        self._month_col: dict[str, int] = {}
        for cell in self.ws[HEADER_ROW]:
            mm = _month_to_mm(cell.value)
            if mm:
                self._month_col[mm] = cell.column

        # Build apartment row map: apt_int → row index
        self._apt_row: dict[int, int] = {}
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            val = self.ws.cell(row=row, column=APT_COL).value
            try:
                apt = int(val)
                self._apt_row[apt] = row
            except (TypeError, ValueError):
                pass

        print(
            f"✓ Ledger ready: {len(self._month_col)} month cols, "
            f"{len(self._apt_row)} apartment rows"
        )

    def write_payment(
        self, apartment: int, month: str, amount: float, overwrite: bool = False
    ) -> tuple[bool, str]:
        """
        Write amount to the cell for (apartment, month).
        month accepts: "04/2026", "אפריל", "04", 4, datetime …
        Returns (success, message).
        """
        mm  = _month_to_mm(month)
        row = self._apt_row.get(int(apartment))
        col = self._month_col.get(mm) if mm else None

        if row is None:
            return False, f"Apt {apartment} not found in ledger"
        if col is None:
            return False, f"Month '{month}' not found in ledger"
        if row > LAST_DATA_ROW:
            return False, f"Row {row} is in the formula section — refused"

        cell = self.ws.cell(row=row, column=col)

        if _is_formula(cell):
            return False, f"{cell.coordinate}: formula — skipped"

        existing = cell.value
        if existing not in (None, "", 0) and not overwrite:
            return False, (
                f"{cell.coordinate}: already has {existing} — "
                f"skipped (pass overwrite=True to replace)"
            )

        cell.value = amount
        col_letter = get_column_letter(col)
        return True, f"✓ {col_letter}{row}  apt {apartment}  ₪{amount:.2f}"

    def save(self, output_path: str | None = None) -> str:
        out = Path(output_path) if output_path else self.path
        self.wb.save(str(out))
        print(f"✓ Saved → {out.name}")
        return str(out)

    @property
    def available_months(self) -> list[str]:
        return sorted(self._month_col.keys())

    @property
    def available_apts(self) -> list[int]:
        return sorted(self._apt_row.keys())


def apply_matches(
    matched: list[dict],
    ledger_path: str,
    payment_month: str,
    overwrite: bool = False,
) -> tuple[list[dict], list[dict]]:
    """
    Open ledger, write all matched payments, save.
    Returns (written, skipped).
    """
    updater = LedgerUpdater(ledger_path)
    written, skipped = [], []

    for result in matched:
        tenant = result.get("tenant")
        if not tenant:
            skipped.append({**result, "write_message": "no tenant resolved"})
            continue

        amount  = result["transaction"]["amount"]
        success, msg = updater.write_payment(
            tenant["apartment"], payment_month, amount, overwrite=overwrite
        )
        entry = {**result, "write_message": msg}
        (written if success else skipped).append(entry)

    updater.save()
    print(f"Written: {len(written)}  |  Skipped: {len(skipped)}")
    return written, skipped


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "האגמית7_כספים_2026.xlsx"
    u = LedgerUpdater(path)
    print("Months:", u.available_months)
    print("Apts  :", u.available_apts[:10], "…")
